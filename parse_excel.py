#!/usr/bin/env python3
"""
SAP PM Weekplanning Excel Parser
Handles both old (W19-25, no Reden) and new (W26+, with Reden) structures.
Exports data suitable for Supabase import.
"""

from openpyxl import load_workbook
from collections import Counter
import json
import sys

WEEKS = [f"Week {i}" for i in range(19, 38)]  # W19-W37

def parse_week(wb, sheet_name):
    """
    Parse a single week sheet.
    Returns dict with week data and metadata.
    """
    try:
        ws = wb[sheet_name]
    except KeyError:
        return None
    
    week_num = int(sheet_name.split()[-1])
    
    # Find header row (contains "Werkzaamheden")
    header_row = None
    for row_num, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
        if row and row[0] == "Werkzaamheden":
            header_row = row_num
            break
    
    if not header_row:
        return {"week": week_num, "rows": [], "has_reden": False, "row_count": 0}
    
    # Check if this week has "Reden" column (W26+)
    headers = tuple(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    has_reden = "Reden" in headers
    
    # Extract data rows
    rows = []
    reasons = []
    
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # Must have at least a task name and order
        if not row or not row[0]:
            continue
        
        task = row[0]
        
        # Skip metadata/headers
        if task in ["Werkzaamheden", "Order", "Werkzaamheden (vulwerk)"] or isinstance(task, str) and task.startswith("="):
            continue
        
        # Skip rows with no order (unless it's a valid task)
        order = row[1]
        if not order and task.strip() in ["2e Pinksterdag", "Geen TD aanwezig", "Hemelvaartsdag"]:
            continue
        
        # Parse columns
        executor1 = row[2] if len(row) > 2 else None
        executor2 = row[3] if len(row) > 3 else None
        executor3 = row[4] if len(row) > 4 else None
        completed = row[5] if len(row) > 5 else None
        
        # Handle Reden/Opmerking based on structure
        if has_reden:
            reden = row[6] if len(row) > 6 else None
            opmerking = row[7] if len(row) > 7 else None
        else:
            reden = None
            opmerking = row[6] if len(row) > 6 else None
        
        # Normalize order (convert to int if possible)
        try:
            order = int(order) if order else None
        except (ValueError, TypeError):
            order = str(order) if order else None
        
        # Build row object
        data_row = {
            "week": week_num,
            "werkzaamheden": str(task).strip() if task else None,
            "order": order,
            "executor1": executor1,
            "executor2": executor2,
            "executor3": executor3,
            "uitgevoerd": str(completed).strip().lower() if completed else None,
            "reden": str(reden).strip() if reden and isinstance(reden, str) else None,
            "opmerking": str(opmerking).strip() if opmerking else None,
        }
        
        rows.append(data_row)
        
        # Track reasons for statistics
        if data_row["reden"]:
            reasons.append(data_row["reden"])
    
    return {
        "week": week_num,
        "rows": rows,
        "has_reden": has_reden,
        "row_count": len(rows),
        "reasons": reasons
    }


def parse_excel(filepath):
    """
    Main parser: reads Excel, returns structured data + statistics.
    """
    wb = load_workbook(filepath, read_only=True)
    
    results = {
        "weeks": [],
        "total_rows": 0,
        "all_reasons": [],
        "reason_summary": {}
    }
    
    for week_name in WEEKS:
        week_data = parse_week(wb, week_name)
        if week_data and week_data["row_count"] > 0:
            results["weeks"].append(week_data)
            results["total_rows"] += week_data["row_count"]
            results["all_reasons"].extend(week_data["reasons"])
    
    # Calculate reason statistics
    if results["all_reasons"]:
        reason_counts = Counter(results["all_reasons"])
        results["reason_summary"] = {
            "total_not_executed": len(results["all_reasons"]),
            "unique_reasons": len(reason_counts),
            "top_reasons": dict(reason_counts.most_common(10))
        }
    
    return results


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 parse_excel.py <filepath>")
        sys.exit(1)
    
    filepath = sys.argv[1]
    data = parse_excel(filepath)
    
    # Pretty print summary
    print("=" * 80)
    print("WEEKPLANNING IMPORT ANALYSIS")
    print("=" * 80)
    
    for week in data["weeks"]:
        structure = "NEW (Reden)" if week["has_reden"] else "OLD"
        print(f"Week {week['week']:2d} [{structure}]: {week['row_count']} rows")
    
    print(f"\nTOTAL: {data['total_rows']} data rows")
    
    if data["reason_summary"]:
        print(f"\nREASONS NOT EXECUTED: {data['reason_summary']['total_not_executed']} instances")
        print("Top reasons:")
        for reason, count in data["reason_summary"]["top_reasons"].items():
            print(f"  {count:3d}x  {reason}")
    
    print("\n" + "=" * 80)
    print("\nJSON output ready for import.")
